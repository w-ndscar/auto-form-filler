#!/usr/bin/env python

from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import keyboard
import openpyxl
from openpyxl.utils import get_column_letter
import customtkinter as ctk
from CTkMessagebox import CTkMessagebox
import time

# This script is used to automate the process of filling a web form with data from an Excel sheet.

def only_integers(char):
    """Function to validate that input is an integer or empty."""
    return char.isdigit() or char == ""

# Launching Chrome
def launch_chrome():
    global driver
    options = Options()
    options.add_experimental_option("detach", True)
    options.add_argument("--start-maximized")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-extensions")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
    driver.get("https://projectmanagement-8d5c4.web.app/")
    #driver.maximize_window()
    CTkMessagebox(title="Info", message="Chrome opened. Please Log in -> Select the Project -> Click on Drawings Tab. Check Instructions for more info")
    time.sleep(2)  # Wait a bit for Chrome to start

def browse_file():
    filename = ctk.filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    file_var.set(filename)

def show_instructions():
    instructions_window = ctk.CTkToplevel(app)
    instructions_window.title("Instructions")
    instructions_window.geometry("720x420")
    instructions_text = (
        "Pre-run check: Select the correct project and ensure all elements are present.\n"
        "If any are missing, create them before running the automation — otherwise, the 'Elements' section will appear empty.\n\n"
        "1. Click 'Open Chrome' and log in to the website\n"
        "2. Select the project and go to the Drawings tab\n"
        "3. Use 'Browse' to select your Excel file\n"
        "4. Enter the sheet name and ending row number\n"
        "5. Click 'Start Automation' to begin\n"
        "\nTip: After entering the values, it'll wait \n"
        "Press 'Esc' key after each entry to proceed to the next \n"
        "DO NOT click Submit/New Buttons. It will be done automatically\n"
        "\n\n\n"
        "Created with curiosity by Arun"
    )
    #ctk.CTkLabel(instructions_window, text="How to Use", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(15, 10))
    ctk.CTkTextbox(instructions_window, width=650, height=370, font=ctk.CTkFont(size=16), border_color="#d3eef7").pack(padx=20, pady=5)
    textbox = instructions_window.winfo_children()[-1]
    textbox.insert("1.0", instructions_text)
    textbox.configure(state="disabled")

def load_excel(file_path, sheet_name, end_row):
    try:
        print("Load excel function called")
        total_rows = int(end_row)
        
        element = []
        sh_size = []
        st_date = []
        dwg_name = []
        dwg_desc = []
        revision = []
        
        # Read data from the specified range
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        print("Reached the Excel file reading part")
        def read_column(sheet, col, start_row, total_rows, date_format=None):
            values = []
            for row in range(start_row, total_rows):
                char = get_column_letter(col)
                cell_name = char + str(row)
                cell_value = sheet[cell_name].value
                if date_format and cell_value:
                    try:
                        cell_value = cell_value.strftime(date_format)
                    except AttributeError:
                        pass # Not a date, keep as is
                values.append(cell_value)
            return values

        start_row = 3

        element = read_column(sheet, 2, start_row, total_rows)
        sh_size = read_column(sheet, 3, start_row, total_rows)
        st_date = read_column(sheet, 4, start_row, total_rows, date_format='%d-%m-%Y')
        dwg_name = read_column(sheet, 5, start_row, total_rows)
        dwg_desc = read_column(sheet, 6, start_row, total_rows)
        revision = read_column(sheet, 7, start_row, total_rows)
        print("Data read from Excel file")

        data = {
            'element': element,
            'sh_size': sh_size,
            'st_date': st_date,
            'dwg_name': dwg_name,
            'dwg_desc': dwg_desc,
            'revision': revision
        }

        return data

    except Exception as e:
        CTkMessagebox(title="Error", message=f"Failed to read Excel file: {e}", icon="cancel")
        return None

def start_automation(data):

    global driver
    
    wait = WebDriverWait(driver, 10)

    element = data['element']
    sh_size = data['sh_size']
    st_date = data['st_date']
    dwg_name = data['dwg_name']
    dwg_desc = data['dwg_desc']
    revision = data['revision']
    print("Starting automation with the following data:")

    def fill_input(by, selector, value, clear=True):
        elem = wait.until(EC.presence_of_element_located((by, selector)))
        if clear:
            elem.clear()
        elem.send_keys(value)

    def select_ng_autocomplete(input_xpath, value, wait):
        # 1. Find the input and type the value
        input_elem = wait.until(EC.element_to_be_clickable((By.XPATH, input_xpath)))
        input_elem.clear()
        input_elem.send_keys(value)
        sleep(1)  # Wait for dropdown to populate (adjust as needed)

        # 2. Press DOWN and ENTER to select the first matching option
        input_elem.send_keys(Keys.ARROW_DOWN)
        input_elem.send_keys(Keys.ENTER)
        sleep(0.5)

    for i in range(len(element)):
        #Click - New Button
        elem2 = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div/div[2]/div/button[1]')))
        elem2.click()
        sleep(1)
        
        print("\n")
        print(" | ")

        #Fields

        #Element
        print (element[i], " | ")
        sleep(0.5)
        # fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/form/div[1]/ng-autocomplete/div[1]/div[1]/input', element[i])
        select_ng_autocomplete('/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/form/div[1]/ng-autocomplete/div[1]/div[1]/input', element[i], wait)

        #Sheet Size - Dropdown
        print (sh_size[i], " | ")
        sleep(0.5)
        Select(driver.find_element(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/form/div[2]/select')).select_by_visible_text(sh_size[i])
        sleep(0.5)

        #Scheduled Date
        print (st_date[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/form/div[3]/input', st_date[i])

        #Drawing name/number
        print (dwg_name[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/form/div[4]/input', dwg_name[i])

        #Drawing Description
        print (dwg_desc[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/form/div[5]/input', dwg_desc[i])
        sleep(0.5)
        
        #Revision
        print (revision[i], " | ")
        fill_input(By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/form/div[6]/input', revision[i])

        #Hotkey to wait before proceeding to Submit
        keyboard.wait("esc")

        #Click - Submit
        elem2 = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/app-root/div/app-project-details/div/div[3]/app-drawings/div[2]/div[2]/div/div/div/div/button[2]')))
        elem2.click()
        sleep(0.5)

#UI Layout 
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")
app = ctk.CTk()
app.title("Add Drawings - Automation")
app.geometry("650x350")

frame_1 = ctk.CTkFrame(app)
frame_1.pack(side="right", pady=20, padx=20, fill="both", expand=True)

frame_2 = ctk.CTkFrame(app)
frame_2.pack(side="left",pady=20, padx=20, fill="both", expand=True)

# Select Excel File
ctk.CTkLabel(frame_2, text="Select Excel:").pack(anchor="center", pady=(15, 0))
file_var = ctk.StringVar()
ctk.CTkEntry(frame_2, textvariable=file_var, width=250).pack(padx=(0), pady=0)
ctk.CTkButton(frame_2, text="Browse", command=browse_file, width=60).pack(pady=5)
#frame.pack_propagate(False)  # Prevent frame from resizing

# Sheet Name and Row Number
sheet_name_var = ctk.StringVar()
ctk.CTkLabel(frame_2, text="Enter Sheet Name:").pack(anchor="center", pady=(15, 0))
ctk.CTkEntry(frame_2, textvariable=sheet_name_var, width=150).pack(pady=0)

row_var = ctk.StringVar()
ctk.CTkLabel(frame_2, text="Ending Row Number:").pack(anchor="center", pady=5)
row_var.set("3")  # Default value
vcmd = (frame_2.register(only_integers), '%P')
ctk.CTkEntry(frame_2, textvariable=row_var, validate='key', validatecommand=vcmd, width=150).pack(pady=5)

# Show Instructions Button
ctk.CTkButton(frame_1, text="Show Instructions", command=show_instructions, fg_color="#188411", border_color="#d3eef7", border_width=2).pack(anchor="center", pady=20)

#Chrome Button
ctk.CTkButton(frame_1, text="Open Chrome", command=launch_chrome, border_color="#d3eef7", border_width=2).pack(anchor="center", pady=30)

# Start Automation Button
ctk.CTkButton(frame_1, text="Start Automation", command=lambda: start_button_clicked(), fg_color="#0a6488", border_color="#d3eef7", border_width=2).pack(anchor="center", pady=40)


def start_button_clicked():
    try:
        if driver is None:
            CTkMessagebox(title="Warning", message="Please open Chrome first.", icon="warning")
            return
        start_automation(load_excel(file_var.get(), sheet_name_var.get(), row_var.get()))
        CTkMessagebox(title="Success", message="Automation completed successfully!", icon="check")
    except Exception as e:
        CTkMessagebox(title="Error", message=f"An error occurred: {e}\nMake sure to open Chrome first", icon="cancel")

app.mainloop()